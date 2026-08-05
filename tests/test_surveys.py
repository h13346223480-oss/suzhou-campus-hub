from io import BytesIO

from openpyxl import load_workbook

from app.extensions import db
from app.models import Survey, SurveyOption, SurveyQuestion, SurveyResponse, User


def make_survey(app, status="published", require_login=False, allow_anonymous=True, slug="test-survey"):
    with app.app_context():
        admin = User.query.filter_by(role="admin").one()
        survey = Survey(title="演示需求调查", slug=slug, description="这是一份用于自动化验收的校园需求调查。",
            status=status, allow_anonymous=allow_anonymous, require_login=require_login,
            estimated_minutes=2, success_message="感谢参与测试调查。", created_by=admin.id)
        db.session.add(survey)
        db.session.flush()
        single = SurveyQuestion(survey_id=survey.id, title="请选择专业方向", question_type="single_choice",
            is_required=True, sort_order=1, validation_rules_json="{}")
        multi = SurveyQuestion(survey_id=survey.id, title="请选择关注内容", question_type="multiple_choice",
            is_required=False, sort_order=2, validation_rules_json='{"max_choices": 2}')
        db.session.add_all([single, multi])
        db.session.flush()
        db.session.add_all([
            SurveyOption(question_id=single.id, label="机器人工程", value="robotics_engineering", sort_order=1),
            SurveyOption(question_id=single.id, label="新能源科学与工程", value="new_energy_science_engineering", sort_order=2),
            SurveyOption(question_id=multi.id, label="宿舍", value="宿舍", sort_order=1),
            SurveyOption(question_id=multi.id, label="地图", value="地图", sort_order=2),
            SurveyOption(question_id=multi.id, label="全英课堂", value="全英课堂", sort_order=3),
        ])
        db.session.commit()
        return survey.id, single.id, multi.id


def test_admin_can_create_survey(client, login, app):
    login("admin@example.com")
    response = client.post("/admin/surveys/create", data={
        "title": "新生真实需求调查", "slug": "freshman-real-needs",
        "description": "用于了解新生真正关心的校园学习和生活问题。", "allow_anonymous": "y",
        "estimated_minutes": 2, "success_message": "感谢参与。",
    }, follow_redirects=True)
    assert "调查草稿已创建" in response.text
    with app.app_context():
        assert Survey.query.filter_by(slug="freshman-real-needs", status="draft").count() == 1


def test_normal_user_cannot_enter_survey_admin(client, login):
    login("verified@example.com")
    assert client.get("/admin/surveys").status_code == 403


def test_published_survey_is_public(client, app):
    make_survey(app)
    response = client.get("/s/test-survey")
    assert response.status_code == 200
    assert "演示需求调查" in response.text
    assert "提交调查" in response.text


def test_draft_survey_is_not_public(client, app):
    make_survey(app, status="draft")
    response = client.get("/s/test-survey")
    assert response.status_code == 404
    assert "提交调查" not in response.text


def test_draft_survey_thanks_page_is_not_public(client, app):
    make_survey(app, status="draft")
    response = client.get("/s/test-survey/thanks")
    assert response.status_code == 404
    assert "演示需求调查" not in response.text


def test_required_question_blocks_submit(client, app):
    make_survey(app)
    client.get("/s/test-survey")
    response = client.post("/s/test-survey", data={"source": "test"}, follow_redirects=True)
    assert "此题为必填题" in response.text
    with app.app_context():
        assert SurveyResponse.query.count() == 0


def test_multiple_choice_maximum_is_enforced(client, app):
    _, single_id, multi_id = make_survey(app)
    client.get("/s/test-survey")
    response = client.post("/s/test-survey", data={
        f"q_{single_id}": "robotics_engineering",
        f"q_{multi_id}": ["宿舍", "地图", "全英课堂"],
    }, follow_redirects=True)
    assert "最多选择 2 项" in response.text
    with app.app_context():
        assert SurveyResponse.query.count() == 0


def test_anonymous_user_can_submit(client, app):
    _, single_id, _ = make_survey(app)
    client.get("/s/test-survey")
    response = client.post("/s/test-survey", data={f"q_{single_id}": "new_energy_science_engineering"}, follow_redirects=True)
    assert "感谢你的参与" in response.text
    with app.app_context():
        record = SurveyResponse.query.one()
        assert record.user_id is None
        assert record.anonymous_token


def test_login_required_survey_redirects(client, app):
    make_survey(app, require_login=True, allow_anonymous=False)
    response = client.get("/s/test-survey")
    assert response.status_code == 302
    assert "/auth/login" in response.headers["Location"]


def test_revoked_student_cannot_submit_login_required_survey(client, login, app):
    make_survey(app, require_login=True, allow_anonymous=False)
    login("pending@example.com")
    response = client.get("/s/test-survey", follow_redirects=True)
    assert "当前账号的校园社区权限不可用" in response.text
    with app.app_context():
        assert SurveyResponse.query.count() == 0


def test_same_account_cannot_submit_twice(client, login, app):
    _, single_id, _ = make_survey(app)
    login("verified@example.com")
    client.post("/s/test-survey", data={f"q_{single_id}": "robotics_engineering"}, follow_redirects=True)
    response = client.post("/s/test-survey", data={f"q_{single_id}": "new_energy_science_engineering"}, follow_redirects=True)
    assert "已经提交过" in response.text
    with app.app_context():
        assert SurveyResponse.query.count() == 1


def test_admin_can_view_statistics(client, login, app):
    survey_id, _, _ = make_survey(app)
    login("admin@example.com")
    response = client.get(f"/admin/surveys/{survey_id}/stats")
    assert response.status_code == 200
    assert "调查统计" in response.text


def test_normal_user_cannot_export(client, login, app):
    survey_id, _, _ = make_survey(app)
    login("verified@example.com")
    assert client.get(f"/admin/surveys/{survey_id}/export.csv").status_code == 403


def test_csv_export(client, login, app):
    survey_id, single_id, _ = make_survey(app)
    client.get("/s/test-survey")
    client.post("/s/test-survey", data={f"q_{single_id}": "robotics_engineering"})
    login("admin@example.com")
    response = client.get(f"/admin/surveys/{survey_id}/export.csv")
    assert response.status_code == 200
    assert response.data.startswith(b"\xef\xbb\xbf")
    assert "请选择专业方向" in response.data.decode("utf-8-sig")


def test_xlsx_export(client, login, app):
    survey_id, single_id, _ = make_survey(app)
    client.get("/s/test-survey")
    client.post("/s/test-survey", data={f"q_{single_id}": "new_energy_science_engineering"})
    login("admin@example.com")
    response = client.get(f"/admin/surveys/{survey_id}/export.xlsx")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    assert workbook.active["A1"].value == "答卷编号"
    assert workbook.active.cell(1, 6).value == "请选择专业方向"


def test_closed_survey_cannot_submit(client, app):
    _, single_id, _ = make_survey(app, status="closed")
    response = client.post("/s/test-survey", data={f"q_{single_id}": "new_energy_science_engineering"}, follow_redirects=True)
    assert "本次调查已结束" in response.text
    with app.app_context():
        assert SurveyResponse.query.count() == 0


def test_source_parameter_is_recorded(client, app):
    _, single_id, _ = make_survey(app)
    client.get("/s/test-survey?source=wechat_group")
    client.post("/s/test-survey?source=wechat_group", data={f"q_{single_id}": "robotics_engineering"})
    with app.app_context():
        assert SurveyResponse.query.one().source == "wechat_group"


def test_qrcode_download_is_png(client, login, app):
    survey_id, _, _ = make_survey(app)
    login("admin@example.com")
    response = client.get(f"/admin/surveys/{survey_id}/qrcode.png")
    assert response.status_code == 200
    assert response.mimetype == "image/png"
    assert response.data.startswith(b"\x89PNG")


def test_anonymous_survey_cookie_is_secure_when_https_cookies_are_enabled(client, app):
    make_survey(app)
    app.config["SESSION_COOKIE_SECURE"] = True
    response = client.get("/s/test-survey")
    cookie = response.headers.get("Set-Cookie", "")
    assert "survey_anonymous_token=" in cookie
    assert "Secure" in cookie
    assert "HttpOnly" in cookie
    assert "SameSite=Lax" in cookie
