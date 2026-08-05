import csv
import json
import math
import re
from collections import Counter
from datetime import datetime, time, timezone
from io import BytesIO, StringIO
from statistics import median
from zoneinfo import ZoneInfo

import qrcode
from flask import (Blueprint, abort, current_app, flash, make_response, redirect, render_template,
                   request, send_file, url_for)
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app.extensions import db
from app.forms import SurveyForm, SurveyQuestionForm
from app.majors import USER_MAJOR_CODES, major_label, normalize_user_major
from app.models import (Survey, SurveyAccessLog, SurveyAnswer, SurveyAnswerTag, SurveyDecisionOverride,
                        SurveyOption, SurveyQuestion, SurveyResponse, SurveyResponseAudit, User)
from app.services.surveys import (answer_display, build_rules, response_audit_details, rules_for,
                                  validate_question_definition)
from app.utils.security import admin_required, contains_html

bp = Blueprint("survey_admin", __name__, url_prefix="/admin/surveys")
SLUG_PATTERN = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
VALIDITY_STATUSES = {"valid", "invalid", "test"}
PRIORITY_LEVELS = {"high", "medium", "watch"}
LOCAL_TIMEZONE = ZoneInfo("Asia/Shanghai")


@bp.route("")
@admin_required
def index():
    surveys = Survey.query.order_by(Survey.updated_at.desc()).all()
    metrics = {}
    for survey in surveys:
        views = SurveyAccessLog.query.filter_by(survey_id=survey.id).count()
        submissions = SurveyResponse.query.filter_by(survey_id=survey.id, validity_status="valid").count()
        metrics[survey.id] = {
            "views": views,
            "submissions": submissions,
            "completion_rate": round(submissions / views * 100, 1) if views else 0,
        }
    return render_template("survey_admin/index.html", surveys=surveys, metrics=metrics)


@bp.route("/create", methods=["GET", "POST"])
@admin_required
def create():
    form = SurveyForm()
    if form.validate_on_submit() and validate_survey_form(form):
        survey = Survey(created_by=current_admin_id())
        apply_survey_form(survey, form)
        db.session.add(survey)
        db.session.commit()
        flash("调查草稿已创建，现在可以添加问题。", "success")
        return redirect(url_for("survey_admin.questions", survey_id=survey.id))
    return render_template("survey_admin/form.html", form=form, survey=None)


@bp.route("/<int:survey_id>/edit", methods=["GET", "POST"])
@admin_required
def edit(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    form = SurveyForm(obj=survey)
    if form.validate_on_submit() and validate_survey_form(form, survey):
        apply_survey_form(survey, form)
        db.session.commit()
        flash("调查基本信息已更新。", "success")
        return redirect(url_for("survey_admin.index"))
    return render_template("survey_admin/form.html", form=form, survey=survey)


def validate_survey_form(form, current=None):
    valid = True
    slug = form.slug.data.strip().lower()
    if not SLUG_PATTERN.match(slug):
        form.slug.errors.append("链接别名只能使用小写字母、数字和连字符。")
        valid = False
    existing = Survey.query.filter_by(slug=slug).first()
    if existing and (not current or existing.id != current.id):
        form.slug.errors.append("该链接别名已存在。")
        valid = False
    if form.start_at.data and form.end_at.data and form.start_at.data >= form.end_at.data:
        form.end_at.errors.append("截止时间必须晚于开始时间。")
        valid = False
    if form.require_login.data:
        form.allow_anonymous.data = False
    return valid


def apply_survey_form(survey, form):
    for name in ["title", "description", "allow_anonymous", "require_login", "allow_edit", "allow_repeat",
                 "use_account_profile_data", "start_at", "end_at", "estimated_minutes", "success_message"]:
        setattr(survey, name, getattr(form, name).data)
    survey.slug = form.slug.data.strip().lower()


def current_admin_id():
    from flask_login import current_user
    return current_user.id


@bp.route("/<int:survey_id>/questions")
@admin_required
def questions(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    return render_template("survey_admin/questions.html", survey=survey)


@bp.route("/<int:survey_id>/questions/add", methods=["GET", "POST"])
@admin_required
def add_question(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    form = SurveyQuestionForm()
    if form.validate_on_submit():
        definition_errors, options = validate_question_definition(form)
        if not definition_errors:
            question = SurveyQuestion(survey_id=survey.id, sort_order=len(survey.questions) + 1)
            apply_question_form(question, form)
            db.session.add(question)
            db.session.flush()
            save_options(question, options, form.add_other.data)
            db.session.commit()
            flash("问题已添加。", "success")
            return redirect(url_for("survey_admin.questions", survey_id=survey.id))
        for error in definition_errors:
            flash(error, "danger")
    return render_template("survey_admin/question_form.html", form=form, survey=survey, question=None)


@bp.route("/<int:survey_id>/questions/<int:question_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_question(survey_id, question_id):
    survey = db.get_or_404(Survey, survey_id)
    question = SurveyQuestion.query.filter_by(id=question_id, survey_id=survey.id).first_or_404()
    form = SurveyQuestionForm(obj=question)
    if request.method == "GET":
        rule_data = rules_for(question)
        form.options_text.data = "\n".join(option.label for option in question.options if option.value != "other")
        form.matrix_rows_text.data = "\n".join(rule_data.get("rows", []))
        form.add_other.data = rule_data.get("allow_other", False)
        for name in ["min_choices", "max_choices", "min_length", "max_length", "min_value", "max_value"]:
            getattr(form, name).data = rule_data.get(name)
    if form.validate_on_submit():
        definition_errors, options = validate_question_definition(form)
        if not definition_errors:
            apply_question_form(question, form)
            question.options.clear()
            db.session.flush()
            save_options(question, options, form.add_other.data)
            db.session.commit()
            flash("问题已更新。", "success")
            return redirect(url_for("survey_admin.questions", survey_id=survey.id))
        for error in definition_errors:
            flash(error, "danger")
    return render_template("survey_admin/question_form.html", form=form, survey=survey, question=question)


def apply_question_form(question, form):
    question.title = form.title.data.strip()
    question.description = (form.description.data or "").strip()
    question.question_type = form.question_type.data
    question.is_required = form.is_required.data
    question.is_contact_info = form.is_contact_info.data
    question.validation_rules_json = json.dumps(build_rules(form), ensure_ascii=False)


def save_options(question, labels, add_other):
    for index, label in enumerate(labels, 1):
        value = "other" if label in {"其他", "其他，请填写"} else label
        db.session.add(SurveyOption(question_id=question.id, label=label, value=value, sort_order=index))
    if add_other and not any(label in {"其他", "其他，请填写"} for label in labels):
        db.session.add(SurveyOption(question_id=question.id, label="其他，请填写", value="other", sort_order=len(labels) + 1))


@bp.route("/<int:survey_id>/questions/<int:question_id>/move/<direction>", methods=["POST"])
@admin_required
def move_question(survey_id, question_id, direction):
    survey = db.get_or_404(Survey, survey_id)
    question = SurveyQuestion.query.filter_by(id=question_id, survey_id=survey.id).first_or_404()
    ordered = list(survey.questions)
    index = ordered.index(question)
    target_index = index - 1 if direction == "up" else index + 1 if direction == "down" else index
    if 0 <= target_index < len(ordered) and target_index != index:
        other = ordered[target_index]
        question.sort_order, other.sort_order = other.sort_order, question.sort_order
        db.session.commit()
    return redirect(url_for("survey_admin.questions", survey_id=survey.id))


@bp.route("/<int:survey_id>/questions/<int:question_id>/delete", methods=["POST"])
@admin_required
def delete_question(survey_id, question_id):
    survey = db.get_or_404(Survey, survey_id)
    question = SurveyQuestion.query.filter_by(id=question_id, survey_id=survey.id).first_or_404()
    if question.answers:
        flash("已有答卷引用该问题，不能删除。可以复制调查后调整。", "warning")
    else:
        SurveyDecisionOverride.query.filter_by(question_id=question.id).delete(
            synchronize_session=False
        )
        db.session.delete(question)
        db.session.flush()
        for index, item in enumerate(survey.questions, 1):
            item.sort_order = index
        db.session.commit()
        flash("问题已删除。", "success")
    return redirect(url_for("survey_admin.questions", survey_id=survey.id))


@bp.route("/<int:survey_id>/copy", methods=["POST"])
@admin_required
def copy_survey(survey_id):
    source = db.get_or_404(Survey, survey_id)
    slug = unique_copy_slug(source.slug)
    copied = Survey(title=f"{source.title}（副本）", slug=slug, description=source.description, status="draft",
        allow_anonymous=source.allow_anonymous, require_login=source.require_login, allow_edit=source.allow_edit,
        allow_repeat=source.allow_repeat, use_account_profile_data=source.use_account_profile_data,
        estimated_minutes=source.estimated_minutes, success_message=source.success_message, created_by=current_admin_id())
    db.session.add(copied)
    db.session.flush()
    for question in source.questions:
        new_question = SurveyQuestion(survey_id=copied.id, title=question.title, description=question.description,
            question_type=question.question_type, is_required=question.is_required, sort_order=question.sort_order,
            is_contact_info=question.is_contact_info, validation_rules_json=question.validation_rules_json)
        db.session.add(new_question)
        db.session.flush()
        for option in question.options:
            db.session.add(SurveyOption(question_id=new_question.id, label=option.label, value=option.value, sort_order=option.sort_order))
    db.session.commit()
    flash("调查已复制为草稿。", "success")
    return redirect(url_for("survey_admin.questions", survey_id=copied.id))


def unique_copy_slug(base):
    counter = 1
    while Survey.query.filter_by(slug=f"{base}-copy-{counter}").first():
        counter += 1
    return f"{base}-copy-{counter}"


@bp.route("/<int:survey_id>/status/<status>", methods=["POST"])
@admin_required
def change_status(survey_id, status):
    survey = db.get_or_404(Survey, survey_id)
    if status not in {"draft", "published", "paused", "closed"}:
        abort(404)
    if status == "published" and not survey.questions:
        flash("至少添加一个问题后才能发布。", "danger")
    else:
        survey.status = status
        db.session.commit()
        flash("调查状态已更新。", "success")
    return redirect(url_for("survey_admin.index"))


@bp.route("/<int:survey_id>/delete", methods=["POST"])
@admin_required
def delete_survey(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    if survey.status != "draft" or survey.responses:
        abort(400)
    SurveyDecisionOverride.query.filter_by(survey_id=survey.id).delete(
        synchronize_session=False
    )
    db.session.delete(survey)
    db.session.commit()
    flash("调查草稿已删除。", "success")
    return redirect(url_for("survey_admin.index"))


@bp.route("/<int:survey_id>/responses")
@admin_required
def responses(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    records = filtered_responses(survey)
    return no_store(render_template("survey_admin/responses.html", survey=survey, responses=records))


@bp.route("/<int:survey_id>/responses/<int:response_id>")
@admin_required
def response_detail(survey_id, response_id):
    survey = db.get_or_404(Survey, survey_id)
    record = SurveyResponse.query.filter_by(id=response_id, survey_id=survey.id).first_or_404()
    answer_map = {answer.question_id: answer for answer in record.answers}
    audit_logs = SurveyResponseAudit.query.filter_by(survey_id=survey.id, response_id=record.id).order_by(
        SurveyResponseAudit.created_at.desc()).all()
    return no_store(render_template("survey_admin/response_detail.html", survey=survey, response=record,
                                    answer_map=answer_map, answer_display=answer_display,
                                    audit_logs=audit_logs))


@bp.route("/<int:survey_id>/responses/<int:response_id>/validity", methods=["POST"])
@admin_required
def update_response_validity(survey_id, response_id):
    survey = db.get_or_404(Survey, survey_id)
    record = SurveyResponse.query.filter_by(id=response_id, survey_id=survey.id).first_or_404()
    status = request.form.get("validity_status", "").strip()
    if status not in VALIDITY_STATUSES:
        abort(400)
    previous = response_validity(record)
    if previous != status:
        record.validity_status = status
        record.is_valid = status == "valid"
        db.session.add(SurveyResponseAudit(
            survey_id=survey.id,
            response_id=record.id,
            actor_id=current_user.id,
            action="validity_changed",
            previous_status=previous,
            new_status=status,
            details_json=response_audit_details(record),
        ))
        db.session.commit()
        flash("答卷有效性已更新，并写入审计记录。", "success")
    return redirect(url_for("survey_admin.response_detail", survey_id=survey.id, response_id=record.id))


@bp.route("/<int:survey_id>/responses/<int:response_id>/delete-confirm")
@admin_required
def confirm_delete_response(survey_id, response_id):
    survey = db.get_or_404(Survey, survey_id)
    record = SurveyResponse.query.filter_by(id=response_id, survey_id=survey.id).first_or_404()
    return no_store(render_template("survey_admin/response_delete_confirm.html", survey=survey, response=record))


@bp.route("/<int:survey_id>/responses/<int:response_id>/delete", methods=["POST"])
@admin_required
def delete_response(survey_id, response_id):
    survey = db.get_or_404(Survey, survey_id)
    record = SurveyResponse.query.filter_by(id=response_id, survey_id=survey.id).first_or_404()
    if request.form.get("confirm_delete") != "yes":
        flash("请在二次确认页勾选确认后再删除。", "danger")
        return redirect(url_for("survey_admin.confirm_delete_response", survey_id=survey.id,
                                response_id=record.id))
    details = json.loads(response_audit_details(record))
    details["deleted_response_id"] = record.id
    details["answer_count"] = len(record.answers)
    for audit in record.audit_logs:
        audit.response_id = None
    db.session.add(SurveyResponseAudit(
        survey_id=survey.id,
        response_id=None,
        actor_id=current_user.id,
        action="deleted",
        previous_status=response_validity(record),
        new_status=None,
        details_json=json.dumps(details, ensure_ascii=False),
    ))
    db.session.delete(record)
    db.session.commit()
    flash("答卷已删除，删除操作和非敏感元数据已写入审计记录。", "success")
    return redirect(url_for("survey_admin.responses", survey_id=survey.id))


@bp.route("/<int:survey_id>/stats")
@admin_required
def stats(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    all_records = filtered_responses(survey)
    valid_records = [record for record in all_records if response_validity(record) == "valid"]
    analysis = analyze_responses(survey, valid_records, total_submissions=len(all_records))
    views = filtered_views(survey)
    analysis["views"] = views
    analysis["completion_rate"] = min(round(len(all_records) / views * 100, 1), 100) if views else 0
    return no_store(render_template("survey_admin/stats.html", survey=survey, analysis=analysis,
                                    filters=request.args))


def filtered_responses(survey, default_validity=None):
    query = SurveyResponse.query.filter_by(survey_id=survey.id)
    date_from = parse_date(request.args.get("date_from"))
    date_to = parse_date(request.args.get("date_to"), end=True)
    if date_from:
        query = query.filter(SurveyResponse.submitted_at >= date_from)
    if date_to:
        query = query.filter(SurveyResponse.submitted_at <= date_to)
    source = request.args.get("source", "").strip()
    if source:
        query = query.filter_by(source=source)
    logged = request.args.get("logged")
    if logged == "yes":
        query = query.filter(SurveyResponse.user_id.is_not(None))
    elif logged == "no":
        query = query.filter(SurveyResponse.user_id.is_(None))
    validity = request.args.get("validity", default_validity or "").strip()
    if validity in VALIDITY_STATUSES:
        query = query.filter(SurveyResponse.validity_status == validity)
    records = query.order_by(SurveyResponse.submitted_at.desc()).all()
    major = request.args.get("major", "").strip()
    year = request.args.get("enrollment_year", type=int)
    if major in USER_MAJOR_CODES:
        records = [record for record in records if response_major_code(record) == major]
    if year:
        records = [record for record in records if record.user and record.user.enrollment_year == year]
    return records


def filtered_views(survey):
    query = SurveyAccessLog.query.filter_by(survey_id=survey.id)
    source = request.args.get("source", "").strip()
    if source:
        query = query.filter_by(source=source)
    date_from = parse_date(request.args.get("date_from"))
    date_to = parse_date(request.args.get("date_to"), end=True)
    if date_from:
        query = query.filter(SurveyAccessLog.visited_at >= date_from)
    if date_to:
        query = query.filter(SurveyAccessLog.visited_at <= date_to)
    return query.count()


def parse_date(value, end=False):
    if not value:
        return None
    try:
        local_day = datetime.strptime(value, "%Y-%m-%d").date()
        local_value = datetime.combine(local_day, time.max if end else time.min, tzinfo=LOCAL_TIMEZONE)
        return local_value.astimezone(timezone.utc)
    except ValueError:
        return None


def analyze_responses(survey, records, total_submissions=None):
    valid_count = len(records)
    daily_counter = Counter(local_datetime(record.submitted_at).strftime("%Y-%m-%d") for record in records)
    daily = dict(sorted(daily_counter.items()))
    sources = Counter(record.source for record in records)
    majors = Counter()
    for record in records:
        code = response_major_code(record)
        majors[major_label(code) if code else "未提供"] += 1
    valid_times = [record.completion_seconds for record in records if record.completion_seconds is not None]
    questions = []
    for question in survey.questions:
        if question.is_contact_info:
            continue
        answers = [answer for record in records for answer in record.answers if answer.question_id == question.id]
        item = {
            "question": question,
            "total": len(answers),
            "option_rows": [],
            "average": None,
            "median": None,
            "distribution": [],
            "high_threshold": None,
            "high_percentage": None,
            "tag_counts": Counter(),
        }
        if question.question_type in {"single_choice", "multiple_choice"}:
            counts = Counter()
            option_labels = {option.value: option.label for option in question.options}
            for answer in answers:
                try:
                    data = json.loads(answer.answer_json or "{}")
                except ValueError:
                    continue
                values = data.get("values", []) if question.question_type == "multiple_choice" else [data.get("value")]
                for value in values:
                    if value:
                        counts[value] += 1
            denominator = valid_count if question.question_type == "multiple_choice" else len(answers)
            rows = []
            known_values = [option.value for option in question.options]
            for value in list(counts):
                if value not in known_values:
                    known_values.append(value)
            for value in known_values:
                count = counts[value]
                rows.append({
                    "value": value,
                    "label": option_labels.get(value, value),
                    "count": count,
                    "percentage": round(count / denominator * 100, 1) if denominator else 0,
                })
            rows.sort(key=lambda row: (-row["count"], known_values.index(row["value"])))
            for rank, row in enumerate(rows, 1):
                row["rank"] = rank
            item["option_rows"] = rows
        elif question.question_type == "rating":
            values = safe_rating_values(answers)
            if values:
                item["average"] = round(sum(values) / len(values), 2)
                item["median"] = round(float(median(values)), 2)
                rules = rules_for(question)
                minimum = float(rules.get("min_value", min(values)))
                maximum = float(rules.get("max_value", max(values)))
                threshold = math.ceil(minimum + (maximum - minimum) * 0.75)
                item["high_threshold"] = threshold
                item["high_percentage"] = round(sum(value >= threshold for value in values) / len(values) * 100, 1)
                distribution = Counter(values)
                if minimum.is_integer() and maximum.is_integer() and 0 <= maximum - minimum <= 20:
                    scale = [float(value) for value in range(int(minimum), int(maximum) + 1)]
                else:
                    scale = sorted(distribution)
                item["distribution"] = [{
                    "label": number_label(value),
                    "count": distribution[value],
                    "percentage": round(distribution[value] / len(values) * 100, 1),
                } for value in scale]
        elif question.question_type in {"short_text", "long_text"}:
            item["tag_counts"] = Counter(tag.tag for answer in answers for tag in answer.tags)
        questions.append(item)
    return {
        "submissions": total_submissions if total_submissions is not None else valid_count,
        "valid_responses": valid_count,
        "average_seconds": round(sum(valid_times) / len(valid_times)) if valid_times else 0,
        "daily": daily,
        "daily_max": max(daily.values(), default=0),
        "sources": sources,
        "majors": majors,
        "questions": questions,
    }


def safe_rating_values(answers):
    values = []
    for answer in answers:
        try:
            values.append(float(answer.answer_text))
        except (TypeError, ValueError):
            continue
    return values


def number_label(value):
    return str(int(value)) if float(value).is_integer() else str(value)


def local_datetime(value):
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(LOCAL_TIMEZONE)


def response_validity(record):
    return record.validity_status or ("valid" if record.is_valid else "invalid")


def response_major_code(record):
    if record.user and record.survey.use_account_profile_data:
        return record.user.major_code
    for answer in record.answers:
        if "专业" in answer.question.title:
            return normalize_user_major(answer_display(answer))
    return None


@bp.route("/<int:survey_id>/text-answers/<int:question_id>")
@admin_required
def text_answers(survey_id, question_id):
    survey = db.get_or_404(Survey, survey_id)
    question = SurveyQuestion.query.filter_by(id=question_id, survey_id=survey.id).first_or_404()
    if question.question_type not in {"short_text", "long_text"}:
        abort(400)
    records = filtered_responses(survey, default_validity="valid")
    answers = [answer for record in records for answer in record.answers if answer.question_id == question.id]
    tag_counts = Counter(tag.tag for answer in answers for tag in answer.tags)
    tag_filter = request.args.get("tag", "").strip()
    if tag_filter:
        answers = [answer for answer in answers if any(tag.tag == tag_filter for tag in answer.tags)]
    return no_store(render_template("survey_admin/text_answers.html", survey=survey, question=question,
                                    answers=answers, tag_counts=tag_counts, tag_filter=tag_filter))


@bp.route("/<int:survey_id>/answers/<int:answer_id>/tags", methods=["POST"])
@admin_required
def update_answer_tags(survey_id, answer_id):
    survey = db.get_or_404(Survey, survey_id)
    answer = SurveyAnswer.query.join(SurveyResponse).filter(
        SurveyAnswer.id == answer_id,
        SurveyResponse.survey_id == survey.id,
    ).first_or_404()
    if answer.question.question_type not in {"short_text", "long_text"}:
        abort(400)
    operation = request.form.get("operation", "add")
    raw_tag = request.form.get("tag", "")
    if operation == "remove":
        tag = SurveyAnswerTag.query.filter_by(answer_id=answer.id, tag=raw_tag.strip()).first_or_404()
        db.session.delete(tag)
        db.session.commit()
        flash("标签已移除。", "success")
    else:
        tags = normalize_tags(raw_tag)
        if not tags:
            flash("请输入 1 至 10 个有效标签，每个标签不超过 40 个字符。", "danger")
        else:
            existing = {item.tag for item in answer.tags}
            for tag in tags:
                if tag not in existing:
                    db.session.add(SurveyAnswerTag(answer_id=answer.id, tag=tag, created_by=current_user.id))
            db.session.commit()
            flash("人工标签已保存。", "success")
    return redirect(url_for("survey_admin.text_answers", survey_id=survey.id,
                            question_id=answer.question_id))


def normalize_tags(raw_value):
    values = [value.strip() for value in re.split(r"[,，]", raw_value or "") if value.strip()]
    if not values or len(values) > 10:
        return []
    result = []
    for value in values:
        if len(value) > 40 or contains_html(value):
            return []
        if value not in result:
            result.append(value)
    return result


@bp.route("/<int:survey_id>/decision-summary", methods=["GET", "POST"])
@admin_required
def decision_summary(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    if request.method == "POST":
        question_id = request.form.get("question_id", type=int)
        option_value = request.form.get("option_value", "").strip()
        priority = request.form.get("priority", "auto")
        question = SurveyQuestion.query.filter_by(id=question_id, survey_id=survey.id).first_or_404()
        if question.is_contact_info or question.question_type not in {"single_choice", "multiple_choice"}:
            abort(400)
        if option_value not in {option.value for option in question.options}:
            abort(400)
        override = SurveyDecisionOverride.query.filter_by(
            survey_id=survey.id, question_id=question.id, option_value=option_value).first()
        if priority == "auto":
            if override:
                db.session.delete(override)
        elif priority in PRIORITY_LEVELS:
            if not override:
                override = SurveyDecisionOverride(survey_id=survey.id, question_id=question.id,
                                                  option_value=option_value, updated_by=current_user.id)
                db.session.add(override)
            override.priority = priority
            override.updated_by = current_user.id
        else:
            abort(400)
        db.session.commit()
        flash("产品决策优先级已更新。", "success")
        return redirect(url_for("survey_admin.decision_summary", survey_id=survey.id))
    records = filtered_responses(survey, default_validity="valid")
    rows = build_decision_rows(survey, records)
    return no_store(render_template("survey_admin/decision_summary.html", survey=survey, rows=rows,
                                    valid_count=len(records)))


def build_decision_rows(survey, records):
    valid_count = len(records)
    analysis = analyze_responses(survey, records)
    overrides = {
        (item.question_id, item.option_value): item.priority
        for item in SurveyDecisionOverride.query.filter_by(survey_id=survey.id).all()
    }
    rows = []
    for item in analysis["questions"]:
        if item["question"].question_type not in {"single_choice", "multiple_choice"}:
            continue
        for option in item["option_rows"]:
            percentage = round(option["count"] / valid_count * 100, 1) if valid_count else 0
            automatic = "high" if percentage > 50 else "medium" if percentage >= 25 else "watch"
            manual = overrides.get((item["question"].id, option["value"]))
            rows.append({
                "question": item["question"],
                "value": option["value"],
                "label": option["label"],
                "count": option["count"],
                "percentage": percentage,
                "automatic_priority": automatic,
                "manual_priority": manual,
                "priority": manual or automatic,
            })
    return sorted(rows, key=lambda row: (-row["percentage"], row["question"].sort_order, row["label"]))


@bp.route("/<int:survey_id>/export.csv")
@admin_required
def export_csv(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    records = filtered_responses(survey)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(export_headers(survey))
    for record in records:
        writer.writerow(export_row(survey, record))
    data = "\ufeff" + output.getvalue()
    return no_store(send_file(BytesIO(data.encode("utf-8")), mimetype="text/csv; charset=utf-8",
                              as_attachment=True, download_name=export_filename(survey, "csv")))


@bp.route("/<int:survey_id>/export.xlsx")
@admin_required
def export_xlsx(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    records = filtered_responses(survey)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "调查答卷"
    sheet.append(export_headers(survey))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173B63")
    for record in records:
        sheet.append(export_row(survey, record))
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 50)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return no_store(send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              as_attachment=True, download_name=export_filename(survey, "xlsx")))


def export_headers(survey):
    return ["答卷编号", "提交时间", "来源", "是否登录用户", "用户专业"] + [
        question.title for question in survey.questions] + ["答卷状态"]


def export_row(survey, record):
    answer_map = {answer.question_id: answer_display(answer) for answer in record.answers}
    major = record.user.major_display if record.user and survey.use_account_profile_data else ""
    return [record.id, local_datetime(record.submitted_at).strftime("%Y-%m-%d %H:%M:%S"),
            record.source, "是" if record.user_id else "否", major] + [
                answer_map.get(question.id, "") for question in survey.questions] + [response_validity(record)]


def export_filename(survey, extension):
    return f"survey_{survey.slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.{extension}"


@bp.route("/<int:survey_id>/anonymous-report.html")
@admin_required
def anonymous_report(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    records = filtered_responses(survey, default_validity="valid")
    analysis = analyze_responses(survey, records)
    html = render_template("survey_admin/anonymous_report.html", survey=survey, analysis=analysis,
                           generated_at=datetime.now(LOCAL_TIMEZONE))
    return no_store(send_file(BytesIO(html.encode("utf-8")), mimetype="text/html; charset=utf-8",
                              as_attachment=True,
                              download_name=f"survey_{survey.slug}_anonymous_summary.html"))


@bp.route("/<int:survey_id>/invite-list.csv")
@admin_required
def export_invite_list(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    contact_questions = [question for question in survey.questions if question.is_contact_info]
    records = filtered_responses(survey, default_validity="valid")
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["答卷编号", "提交时间"] + [question.title for question in contact_questions])
    for record in records:
        answer_map = {answer.question_id: answer_display(answer) for answer in record.answers}
        values = [answer_map.get(question.id, "") for question in contact_questions]
        if any(values):
            writer.writerow([record.id, local_datetime(record.submitted_at).strftime("%Y-%m-%d %H:%M:%S")] + values)
    data = "\ufeff" + output.getvalue()
    return no_store(send_file(BytesIO(data.encode("utf-8")), mimetype="text/csv; charset=utf-8",
                              as_attachment=True,
                              download_name=f"survey_{survey.slug}_invite_list.csv"))


def no_store(response):
    if isinstance(response, str):
        response = make_response(response)
    response.headers["Cache-Control"] = "private, no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@bp.route("/<int:survey_id>/qrcode.png")
@admin_required
def qrcode_png(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    public_url = f"{current_app.config['APP_BASE_URL']}{url_for('surveys.fill', slug=survey.slug)}"
    image = qrcode.make(public_url)
    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return send_file(output, mimetype="image/png", as_attachment=True, download_name=f"survey_{survey.slug}_qrcode.png")


@bp.route("/<int:survey_id>/share")
@admin_required
def share(survey_id):
    survey = db.get_or_404(Survey, survey_id)
    public_url = f"{current_app.config['APP_BASE_URL']}{url_for('surveys.fill', slug=survey.slug)}"
    return render_template("survey_admin/share.html", survey=survey, public_url=public_url)
